import copy
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook

SRC = '/home/user/Downloads/JD6628H_Test Plan_20260416.xlsx'
BASE = '/home/user/workspace/tenji-bot/templates/CASE_2026-04-01_JD6628H_BM2922AA_JD6628H_TenjiV2.2_R0.1.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_SinglePwr_2-10_to_2-18_SplitRows_TenjiV2.2_R0.1.xlsm'

COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}

PDO_TO_VOLT = {'PDO1': 5, 'PDO2': 9, 'PDO3': 12, 'PDO5': 20}


def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def split_lines(v):
    text = clean(v)
    return [x.strip() for x in text.split('\n') if x.strip()]


def merged_children(ws):
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                children.add((row, col))
    return children


def clear_test_note_values(ws, start_row, start_col, end_col):
    child_cells = merged_children(ws)
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))


def copy_row_style(ws, src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
        if src._style:
            dst._style = copy.copy(src._style)


def set_row_height(ws, row_idx):
    max_lines = 1
    for c in range(2, 15):
        val = ws.cell(row_idx, c).value
        if isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)
    ws.row_dimensions[row_idx].height = min(15 * max_lines, 200)


def parse_voltage_token(text):
    m = re.search(r'(PDO\d+)', text)
    if m and m.group(1) in PDO_TO_VOLT:
        return PDO_TO_VOLT[m.group(1)]
    m = re.search(r'(5|9|12|20)V', text)
    if m:
        return int(m.group(1))
    raise ValueError(f'Cannot parse voltage from: {text!r}')


def judge_window(voltage):
    voltage = int(voltage)
    if voltage == 5:
        return '4', '6'
    if voltage == 9:
        return '7.2', '10.8'
    if voltage == 12:
        return '10', '13.5'
    if voltage == 20:
        return '16', '21.5'
    return str(max(0, voltage - 1)), str(voltage + 1)


def pd_req(voltage):
    return f'PD_req{int(voltage)}v'


def step_measure(step_idx, primary_v):
    p_lo, p_hi = judge_window(primary_v)
    if step_idx == 1:
        return '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            f'PD_command {pd_req(primary_v)}',
            'Wait 20ms',
            f'JudgeV VBUSC_C2 {p_lo} {p_hi}',
            'JudgeV VBUSC_C1 0 0.8',
        ])
    if step_idx == 2:
        return '\n'.join([
            'Wait 100ms',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req5v',
            'Wait 30ms',
            'JudgeV VBUSC_C2 4 6',
            'JudgeV VBUSC_C1 4 6',
        ])
    if step_idx == 3:
        return '\n'.join([
            'Wait 100ms',
            'ForceV vatach1 0V 0A',
            'UR k3 OFF',
            'Wait 20ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            f'PD_command {pd_req(primary_v)}',
            'Wait 30ms',
            f'JudgeV VBUSC_C2 {p_lo} {p_hi}',
            'JudgeV VBUSC_C1 0 0.8',
        ])
    raise ValueError(step_idx)


def step_symbol(no, step_idx, primary_v):
    if step_idx == 1:
        return f'SP_{no.replace("-", "_")}_C2_ONLY_{primary_v}V'
    if step_idx == 2:
        return f'SP_{no.replace("-", "_")}_C2_C1_SHARE_5V'
    if step_idx == 3:
        return f'SP_{no.replace("-", "_")}_C2_RECOVER_{primary_v}V'
    raise ValueError(step_idx)


def step_spec(step_idx, primary_v, row_idx):
    typ = 5 if step_idx == 2 else int(primary_v)
    return {
        'min': f'=K{row_idx}*0.8',
        'typ': typ,
        'max': f'=K{row_idx}*1.2',
        'unit': 'V',
    }


def write_os_row(ws, row_idx, symbol='OS_test'):
    ws.cell(row_idx, 1).value = None
    ws.cell(row_idx, COL['bin']).value = 1 if symbol == 'OS_test' else 12
    ws.cell(row_idx, COL['test_item']).value = symbol
    ws.cell(row_idx, COL['symbol']).value = symbol
    ws.cell(row_idx, COL['pwr_seq']).value = 'PWR_OS'
    ws.cell(row_idx, COL['pseudo']).value = None
    ws.cell(row_idx, COL['run']).value = None
    ws.cell(row_idx, COL['measure']).value = None
    ws.cell(row_idx, COL['desc']).value = 'open and short test'
    ws.cell(row_idx, COL['min']).value = -0.3
    ws.cell(row_idx, COL['typ']).value = None
    ws.cell(row_idx, COL['max']).value = -0.8
    ws.cell(row_idx, COL['unit']).value = 'V'
    ws.cell(row_idx, COL['note']).value = None


src_wb = load_workbook(SRC, read_only=True, data_only=False)
src_ws = src_wb['單電源模式']
expanded = []
want = {f'2-{i}' for i in range(10, 19)}
for r in range(16, src_ws.max_row + 1):
    no = clean(src_ws.cell(r, 1).value)
    if no not in want:
        continue
    expected_lines = split_lines(src_ws.cell(r, 5).value)
    target_lines = split_lines(src_ws.cell(r, 7).value)
    request_lines = split_lines(src_ws.cell(r, 9).value)
    primary_v = parse_voltage_token(request_lines[0])
    notes = '\n'.join([
        f'Source row/item: 單電源{no}',
        f'Category: {clean(src_ws.cell(r, 2).value)}',
        f'Function: {clean(src_ws.cell(r, 3).value)}',
        f'Test method: {clean(src_ws.cell(r, 4).value)}',
        f'Expected: {clean(src_ws.cell(r, 5).value)}',
        f'Target voltage: {clean(src_ws.cell(r, 7).value)}',
        f'Request PDO: {clean(src_ws.cell(r, 9).value)}',
        f'Protocol: {clean(src_ws.cell(r, 10).value)}',
        f'Verify: {clean(src_ws.cell(r, 11).value)}',
        'Mapping: split each observable step into one Test Note row so each Judge aligns to one SPEC fragment',
    ])
    for step_idx in [1, 2, 3]:
        expanded.append({
            'no': no,
            'step_idx': step_idx,
            'primary_v': primary_v,
            'symbol': step_symbol(no, step_idx, primary_v),
            'desc': expected_lines[step_idx - 1] if len(expected_lines) >= step_idx else '',
            'target': target_lines[step_idx - 1] if len(target_lines) >= step_idx else '',
            'measure': step_measure(step_idx, primary_v),
            'note': notes,
        })

expanded.sort(key=lambda x: (int(x['no'].split('-')[1]), x['step_idx']))

shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']
clear_test_note_values(ws, start_row=2, start_col=2, end_col=14)
copy_row_style(ws, 2, 2)
write_os_row(ws, 2, 'OS_test')

body_style_rows = {1: 3, 2: 4, 3: 5}
row_idx = 3
for item in expanded:
    copy_row_style(ws, body_style_rows[item['step_idx']], row_idx)
    spec = step_spec(item['step_idx'], item['primary_v'], row_idx)
    ws.cell(row_idx, COL['bin']).value = '5'
    ws.cell(row_idx, COL['test_item']).value = 'SinglePower_2' if item['step_idx'] == 1 else None
    ws.cell(row_idx, COL['symbol']).value = item['symbol']
    ws.cell(row_idx, COL['pwr_seq']).value = 'PWR_VBUS'
    ws.cell(row_idx, COL['pseudo']).value = None
    ws.cell(row_idx, COL['run']).value = None
    ws.cell(row_idx, COL['measure']).value = item['measure']
    ws.cell(row_idx, COL['desc']).value = item['desc']
    ws.cell(row_idx, COL['min']).value = spec['min']
    ws.cell(row_idx, COL['typ']).value = spec['typ']
    ws.cell(row_idx, COL['max']).value = spec['max']
    ws.cell(row_idx, COL['unit']).value = spec['unit']
    ws.cell(row_idx, COL['note']).value = item['note'] + f'\nStep target: {item["target"]}'
    set_row_height(ws, row_idx)
    row_idx += 1

copy_row_style(ws, 2, row_idx)
write_os_row(ws, row_idx, 'OS_test2')
set_row_height(ws, row_idx)

tail_os_row = row_idx
child_cells = merged_children(ws)
for r in range(tail_os_row + 1, max(ws.max_row, tail_os_row + 10) + 1):
    for c in range(1, 15):
        if (r, c) in child_cells:
            continue
        ws.cell(r, c).value = None

if 'Revision history' in wb.sheetnames:
    rev_ws = wb['Revision history']
    rev_row = 3
    while rev_ws.cell(rev_row, 1).value:
        rev_row += 1
    rev_ws.cell(rev_row, 1).value = '2026-04-27'
    rev_ws.cell(rev_row, 2).value = 'R0.1'
    rev_ws.cell(rev_row, 3).value = 'Hermes'
    rev_ws.cell(rev_row, 4).value = 'Generate JD6628H 單電源2-10~2-18 split-row TENJI workbook with workbook head/tail OS_test and per-step SPEC'

wb.save(OUT)
print(f'output={OUT}')
print(f'written_rows=3-{tail_os_row-1}')
print(f'tail_os_row={tail_os_row}')
print('total_rows', len(expanded))
for item in expanded[:6]:
    print(item['symbol'], '=>', item['desc'])
