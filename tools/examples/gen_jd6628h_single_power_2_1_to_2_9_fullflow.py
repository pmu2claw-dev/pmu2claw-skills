import copy
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook

SRC = '/home/user/Downloads/JD6628H_Test Plan_20260416.xlsx'
BASE = '/home/user/workspace/tenji-bot/templates/CASE_2026-04-01_JD6628H_BM2922AA_JD6628H_TenjiV2.2_R0.1.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_SinglePwr_2-1_to_2-9_FullFlow_TenjiV2.2_R0.1.xlsm'

COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}

PDO_TO_VOLT = {
    'PDO1': 5,
    'PDO2': 9,
    'PDO3': 12,
    'PDO5': 20,
}


def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def split_lines(v):
    text = clean(v)
    return [x.strip() for x in text.split('\n') if x.strip()]


def merged_children(ws):
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                children.add((row, col))
    return children


def clear_test_note_values(ws, start_row, start_col, end_col):
    child_cells = merged_children(ws)
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))


def copy_row_style(ws, src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
        if src._style:
            dst._style = copy.copy(src._style)


def set_row_height(ws, row_idx):
    max_lines = 1
    for c in range(2, 15):
        val = ws.cell(row_idx, c).value
        if isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)
    ws.row_dimensions[row_idx].height = min(15 * max_lines, 260)


def parse_voltage_token(text):
    m = re.search(r'(PDO\d+)', text)
    if m and m.group(1) in PDO_TO_VOLT:
        return PDO_TO_VOLT[m.group(1)]
    m = re.search(r'(5|9|12|20)V', text)
    if m:
        return int(m.group(1))
    raise ValueError(f'Cannot parse voltage from: {text!r}')


def judge_window(voltage):
    voltage = int(voltage)
    if voltage == 5:
        return '4', '6'
    if voltage == 9:
        return '7.2', '10.8'
    if voltage == 12:
        return '10', '13.5'
    if voltage == 20:
        return '16', '21.5'
    return str(max(0, voltage - 1)), str(voltage + 1)


def pd_command(voltage):
    return f'PD_req{int(voltage)}v'


def symbol_for(no, primary_v, secondary_v):
    return f'SP_{no.replace("-", "_")}_C1_{primary_v}V_C2_{secondary_v}V'


def build_description(no, primary_v, secondary_v):
    return f'單電源模式 {no}：先插C1({primary_v}V)→插拔C2({secondary_v}V)；C1降為5V共享，拔除C2後回復{primary_v}V'


def build_measure(primary_v, secondary_v):
    p_lo, p_hi = judge_window(primary_v)
    lines = [
        'Wait 10ms',
        'ForceV VIN 5 200mA',
        'ForceV vatach1 7V 0A',
        'UR k3 ON',
        'Wait 1ms',
        'PD_command PD_INIT',
        'Wait 2ms',
        f'PD_command {pd_command(primary_v)}',
        'Wait 20ms',
        f'JudgeV VBUSC_C1 {p_lo} {p_hi}',
        'JudgeV VBUSC_C2 0 0.8',
        'Wait 100ms',
        'ForceV vatach2 7V 0A',
        'UR k4 ON',
        'Wait 1ms',
        'PD_command PD_INIT',
        'Wait 2ms',
        f'PD_command {pd_command(secondary_v)}',
        'Wait 30ms',
        'JudgeV VBUSC_C1 4 6',
        'JudgeV VBUSC_C2 4 6',
        'Wait 100ms',
        'ForceV vatach2 0V 0A',
        'UR k4 OFF',
        'Wait 20ms',
        'PD_command PD_INIT',
        'Wait 2ms',
        f'PD_command {pd_command(primary_v)}',
        'Wait 30ms',
        f'JudgeV VBUSC_C1 {p_lo} {p_hi}',
        'JudgeV VBUSC_C2 0 0.8',
    ]
    return '\n'.join(lines)


src_wb = load_workbook(SRC, read_only=True, data_only=False)
src_ws = src_wb['單電源模式']
rows = []
want = {f'2-{i}' for i in range(1, 10)}
for r in range(16, src_ws.max_row + 1):
    no = clean(src_ws.cell(r, 1).value)
    if no not in want:
        continue
    request_lines = split_lines(src_ws.cell(r, 9).value)
    expected_lines = split_lines(src_ws.cell(r, 5).value)
    primary_v = parse_voltage_token(request_lines[0])
    secondary_v = parse_voltage_token(request_lines[1]) if len(request_lines) > 1 else 5
    rows.append({
        'src_row': r,
        'no': no,
        'category': clean(src_ws.cell(r, 2).value),
        'test_func': clean(src_ws.cell(r, 3).value),
        'test_method': clean(src_ws.cell(r, 4).value),
        'expected': clean(src_ws.cell(r, 5).value),
        'request_pdo': clean(src_ws.cell(r, 9).value),
        'verify': clean(src_ws.cell(r, 11).value),
        'sim': clean(src_ws.cell(r, 12).value),
        'owner': clean(src_ws.cell(r, 13).value),
        'due': clean(src_ws.cell(r, 14).value),
        'primary_v': primary_v,
        'secondary_v': secondary_v,
        'description': build_description(no, primary_v, secondary_v),
        'measure': build_measure(primary_v, secondary_v),
        'symbol': symbol_for(no, primary_v, secondary_v),
        'note': '\n'.join([
            f'Source row/item: 單電源{no}',
            f'Category: {clean(src_ws.cell(r, 2).value)}',
            f'Function: {clean(src_ws.cell(r, 3).value)}',
            f'Request PDO: {clean(src_ws.cell(r, 9).value)}',
            f'Expected: {clean(src_ws.cell(r, 5).value)}',
            f'Verify: {clean(src_ws.cell(r, 11).value)}',
            f'sim: {clean(src_ws.cell(r, 12).value)}',
            f'Owner: {clean(src_ws.cell(r, 13).value)}',
            f'Due: {clean(src_ws.cell(r, 14).value)}',
            'Mapping: practice-like full-flow reconstruction for C1-first single-power family',
        ]),
    })

rows.sort(key=lambda x: int(x['no'].split('-')[1]))

shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']
clear_test_note_values(ws, start_row=2, start_col=2, end_col=14)
style_row = 2
row_idx = 3
for item in rows:
    copy_row_style(ws, style_row, row_idx)
    ws.cell(row_idx, COL['bin']).value = '5'
    ws.cell(row_idx, COL['test_item']).value = 'SinglePower_2'
    ws.cell(row_idx, COL['symbol']).value = item['symbol']
    ws.cell(row_idx, COL['pwr_seq']).value = 'PWR_VBUS'
    ws.cell(row_idx, COL['pseudo']).value = None
    ws.cell(row_idx, COL['run']).value = None
    ws.cell(row_idx, COL['measure']).value = item['measure']
    ws.cell(row_idx, COL['desc']).value = item['description']
    ws.cell(row_idx, COL['min']).value = None
    ws.cell(row_idx, COL['typ']).value = None
    ws.cell(row_idx, COL['max']).value = None
    ws.cell(row_idx, COL['unit']).value = 'V'
    ws.cell(row_idx, COL['note']).value = item['note']
    set_row_height(ws, row_idx)
    row_idx += 1

if 'Revision history' in wb.sheetnames:
    rev_ws = wb['Revision history']
    rev_row = 3
    while rev_ws.cell(rev_row, 1).value:
        rev_row += 1
    rev_ws.cell(rev_row, 1).value = '2026-04-24'
    rev_ws.cell(rev_row, 2).value = 'R0.1'
    rev_ws.cell(rev_row, 3).value = 'Hermes'
    rev_ws.cell(rev_row, 4).value = 'Rebuild JD6628H 單電源2-1~2-9 into practice-like full TENJI flow workbook'

wb.save(OUT)
print(f'output={OUT}')
print(f'source_items={len(rows)}')
print(f'written_rows=3-{row_idx-1}')
for item in rows:
    print(item['no'], item['symbol'])
