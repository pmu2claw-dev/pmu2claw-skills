import copy
import shutil
from datetime import datetime
from openpyxl import load_workbook

SRC = '/home/user/Downloads/JD6628H_Test Plan_20260416.xlsx'
BASE = '/home/user/workspace/tenji-bot/templates/CASE_2026-04-01_JD6628H_BM2922AA_JD6628H_TenjiV2.2_R0.1.xlsm'
REF = '/home/user/Downloads/JD6628H_BM2922AA_Dual2-1_Blind2-3_TenjiV2.2_R0.1.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_BlindInsert_2-3_SplitRows_TenjiV2.2_R0.2.xlsm'

COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}


def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def merged_children(ws):
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                children.add((row, col))
    return children


def clear_test_note_values(ws, start_row, start_col, end_col):
    child_cells = merged_children(ws)
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))


def copy_row_style(src_ws, src_row, dst_ws, dst_row):
    for c in range(1, max(src_ws.max_column, dst_ws.max_column) + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
        if src._style:
            dst._style = copy.copy(src._style)


def set_row_height(ws, row_idx, height=None):
    if height is not None:
        ws.row_dimensions[row_idx].height = height
        return
    max_lines = 1
    for c in range(2, 15):
        val = ws.cell(row_idx, c).value
        if isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)
    ws.row_dimensions[row_idx].height = min(15 * max_lines, 240)


def write_values(ws, row_idx, data):
    ws.cell(row_idx, COL['bin']).value = data['bin']
    ws.cell(row_idx, COL['test_item']).value = data['test_item']
    ws.cell(row_idx, COL['symbol']).value = data['symbol']
    ws.cell(row_idx, COL['pwr_seq']).value = data['pwr_seq']
    ws.cell(row_idx, COL['pseudo']).value = data['pseudo']
    ws.cell(row_idx, COL['run']).value = data['run']
    ws.cell(row_idx, COL['measure']).value = data['measure']
    ws.cell(row_idx, COL['desc']).value = data['desc']
    ws.cell(row_idx, COL['min']).value = data['min']
    ws.cell(row_idx, COL['typ']).value = data['typ']
    ws.cell(row_idx, COL['max']).value = data['max']
    ws.cell(row_idx, COL['unit']).value = data['unit']
    ws.cell(row_idx, COL['note']).value = data['note']


def os_row_dict(ws, row_idx):
    return {
        'bin': ws.cell(row_idx, COL['bin']).value,
        'test_item': ws.cell(row_idx, COL['test_item']).value,
        'symbol': ws.cell(row_idx, COL['symbol']).value,
        'pwr_seq': ws.cell(row_idx, COL['pwr_seq']).value,
        'pseudo': ws.cell(row_idx, COL['pseudo']).value,
        'run': ws.cell(row_idx, COL['run']).value,
        'measure': ws.cell(row_idx, COL['measure']).value,
        'desc': ws.cell(row_idx, COL['desc']).value,
        'min': ws.cell(row_idx, COL['min']).value,
        'typ': ws.cell(row_idx, COL['typ']).value,
        'max': ws.cell(row_idx, COL['max']).value,
        'unit': ws.cell(row_idx, COL['unit']).value,
        'note': ws.cell(row_idx, COL['note']).value,
    }


src_wb = load_workbook(SRC, data_only=False)
src_ws = src_wb['智能盲插模式']
source_row = None
for r in range(16, src_ws.max_row + 1):
    if clean(src_ws.cell(r, 1).value) == '2-3':
        source_row = r
        break
if source_row is None:
    raise SystemExit('source row 2-3 not found')

ref_wb = load_workbook(REF, keep_vba=True, data_only=False)
ref_ws = ref_wb['Test Note']

notes_common = '\n'.join([
    'Source row/item: 智能盲插2-3',
    f'Category: {clean(src_ws.cell(source_row, 2).value)}',
    f'Function: {clean(src_ws.cell(source_row, 3).value)}',
    f'Test method: {clean(src_ws.cell(source_row, 4).value)}',
    f'Expected: {clean(src_ws.cell(source_row, 5).value)}',
    f'Target voltage(G25): {clean(src_ws.cell(source_row, 7).value)}',
    f'Request PDO: {clean(src_ws.cell(source_row, 9).value)}',
    f'Protocol: {clean(src_ws.cell(source_row, 10).value)}',
    f'Verify: {clean(src_ws.cell(source_row, 11).value)}',
    'Answer reference: JD6628H_BM2922AA_Dual2-1_Blind2-3_TenjiV2.2_R0.1.xlsm / Test Note row 4',
    'Repair note: prior R0.1 wrongly collapsed multiple verifiable Judge targets into four rows',
    'Repair rule: split one TENJI row per verifiable Judge target; by G25 voltage fragments alone this case needs at least 8 body rows',
    'Observed source anomaly: G25 step-3/4 voltage wording conflicts with the answer workbook measured-node/result mapping; row split below follows answer-verified C1/C2 Judge targets while preserving source traceability',
])

rows = [
    {
        'bin': '5',
        'test_item': 'BlindInsert_2',
        'symbol': 'BI_2_3_S1_C1_9V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'JudgeV VBUSC_C1 7.2 10.8',
        ]),
        'desc': '1.C1口單插，C1口輸出9V。',
        'min': 7.2,
        'typ': 9,
        'max': 10.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step1 / judge target = C1 9V',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S1_C2_OFF',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'JudgeV VBUSC_C2 0 0.8',
        ]),
        'desc': '1.C1口單插，C2口無輸出。',
        'min': 0,
        'typ': 0,
        'max': 0.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step1 / judge target = C2 absent',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S2_C1_9V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'JudgeV VBUSC_C1 7.2 10.8',
        ]),
        'desc': '2.C2口插入時，C1口維持9V。',
        'min': 7.2,
        'typ': 9,
        'max': 10.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step2 / judge target = C1 9V hold',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S2_C2_5V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'JudgeV VBUSC_C2 4 6',
        ]),
        'desc': '2.C2口插入時，C2口先落在5V fallback。',
        'min': 4,
        'typ': 5,
        'max': 6,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step2 / judge target = C2 5V fallback',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S3_C1_9V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req20v',
            'Wait 30ms',
            'JudgeV VBUSC_C1 7.2 10.8',
        ]),
        'desc': '3.C2口轉主路輸出20V後，C1口轉輔路維持9V。',
        'min': 7.2,
        'typ': 9,
        'max': 10.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step3 / judge target = C1 9V hold',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S3_C2_20V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req20v',
            'Wait 30ms',
            'JudgeV VBUSC_C2 16 21.5',
        ]),
        'desc': '3.C2口轉主路輸出20V。',
        'min': 16,
        'typ': 20,
        'max': 21.5,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step3 / judge target = C2 20V transfer',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S4_C1_9V',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req20v',
            'Wait 30ms',
            'ForceV vatach2 0V 0A',
            'UR k4 OFF',
            'Wait 20ms',
            'JudgeV VBUSC_C1 7.2 10.8',
        ]),
        'desc': '4.C2口拔除後，C1口轉主路維持9V。',
        'min': 7.2,
        'typ': 9,
        'max': 10.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step4 / judge target = C1 recover 9V',
    },
    {
        'bin': '5',
        'test_item': None,
        'symbol': 'BI_2_3_S4_C2_OFF',
        'pwr_seq': 'PWR_VBUS',
        'pseudo': None,
        'run': None,
        'measure': '\n'.join([
            'Wait 10ms',
            'ForceV VIN 5 200mA',
            'ForceV vatach1 7V 0A',
            'UR k3 ON',
            'Wait 1ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req9v',
            'Wait 20ms',
            'ForceV vatach2 7V 0A',
            'UR k4 ON',
            'Wait 20ms',
            'PD_command PD_INIT',
            'Wait 2ms',
            'PD_command PD_req20v',
            'Wait 30ms',
            'ForceV vatach2 0V 0A',
            'UR k4 OFF',
            'Wait 20ms',
            'JudgeV VBUSC_C2 0 0.8',
        ]),
        'desc': '4.C2口拔除後，C2口無輸出。',
        'min': 0,
        'typ': 0,
        'max': 0.8,
        'unit': 'V',
        'note': notes_common + '\nSplit row: step4 / judge target = C2 absent',
    },
]

shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']
clear_test_note_values(ws, start_row=2, start_col=2, end_col=14)

copy_row_style(ref_ws, 2, ws, 2)
write_values(ws, 2, os_row_dict(ref_ws, 2))
set_row_height(ws, 2, ref_ws.row_dimensions[2].height)

for idx, row in enumerate(rows, start=3):
    copy_row_style(ref_ws, 4, ws, idx)
    write_values(ws, idx, row)
    set_row_height(ws, idx)

tail_row = 3 + len(rows)
copy_row_style(ref_ws, 5, ws, tail_row)
write_values(ws, tail_row, os_row_dict(ref_ws, 5))
set_row_height(ws, tail_row, ref_ws.row_dimensions[5].height)

child_cells = merged_children(ws)
for r in range(tail_row + 1, max(ws.max_row, 20) + 1):
    for c in range(1, 15):
        if (r, c) in child_cells:
            continue
        ws.cell(r, c).value = None

if 'Revision history' in wb.sheetnames:
    rev_ws = wb['Revision history']
    rev_row = 3
    while rev_ws.cell(rev_row, 1).value:
        rev_row += 1
    rev_ws.cell(rev_row, 1).value = '2026-04-27'
    rev_ws.cell(rev_row, 2).value = 'R0.2'
    rev_ws.cell(rev_row, 3).value = 'Hermes'
    rev_ws.cell(rev_row, 4).value = 'Repair BlindInsert 2-3 row splitting: expand collapsed 4-row output to 8 body rows, one per verifiable Judge target'

wb.save(OUT)
print(f'output={OUT}')
print(f'body_rows={len(rows)}')
print(f'tail_row={tail_row}')
for r in range(2, tail_row + 1):
    print(r, ws.cell(r, COL['symbol']).value, ws.cell(r, COL['typ']).value)
