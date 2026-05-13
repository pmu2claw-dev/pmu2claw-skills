import copy
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook

SRC = '/home/user/Downloads/JD6628H_Test Plan_20260416.xlsx'
BASE = '/home/user/workspace/tenji-bot/templates/CASE_2026-04-01_JD6628H_BM2922AA_JD6628H_TenjiV2.2_R0.1.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_SinglePwr_NonFPGA_TenjiV2.2_R0.1.xlsm'


EXCLUDE_VERIFY_KEYWORDS = ('FPGA',)


def should_exclude_verify(verify_text):
    verify = clean(verify_text).upper()
    return any(keyword in verify for keyword in EXCLUDE_VERIFY_KEYWORDS)


def merged_children(ws):
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                children.add((row, col))
    return children


def clear_test_note_values(ws, start_row, start_col, end_col):
    child_cells = merged_children(ws)
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))



COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}


def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def split_lines(v):
    text = clean(v)
    if not text:
        return []
    return [x.strip() for x in text.split('\n') if x and x.strip()]


def strip_num_prefix(line):
    return re.sub(r'^\s*\d+[\.-]?\s*', '', clean(line))


def infer_group(no, test_func):
    tf = clean(test_func)
    if no.startswith('1-'):
        return 'SinglePwr_Protocol'
    if no.startswith('2-'):
        if '先插C1，再插拔C2' in tf:
            return 'SinglePwr_C1C2'
        if '先插C2，再插拔C1' in tf:
            return 'SinglePwr_C2C1'
        if '先插C1，再插拔A' in tf:
            return 'SinglePwr_C1A'
        if '先插C2，再插拔A' in tf:
            return 'SinglePwr_C2A'
        if '先插A，再插拔C1' in tf:
            return 'SinglePwr_AC1'
        if '先插A，再插拔C2' in tf:
            return 'SinglePwr_AC2'
        if '先插C1，再插上C2' in tf and '最後拔掉A' in tf:
            return 'SinglePwr_C1C2A'
        if '先插C2，再插上C1' in tf and '最後拔掉A' in tf:
            return 'SinglePwr_C2C1A'
        if '先插C1，再插上A' in tf and '最後拔掉C2' in tf:
            return 'SinglePwr_C1AC2'
        if '先插C2，再插上A' in tf and '最後拔掉C1' in tf:
            return 'SinglePwr_C2AC1'
        return 'SinglePwr_Mode'
    if no.startswith('3-'):
        return 'SinglePwr_Load'
    if no.startswith('4-'):
        return 'SinglePwr_OCP'
    return 'SinglePwr_Other'


def infer_run_pattern(step_line):
    step = clean(step_line)
    if not step:
        return ''
    m = re.search(r'等候\s*(\d+)\s*秒', step)
    if m:
        return f'Wait {m.group(1)}s'
    m = re.search(r'Wait\s*(\d+)\s*s', step, re.I)
    if m:
        return f'Wait {m.group(1)}s'
    return ''


def infer_typ_voltage(expected_line, target_line):
    exp = clean(expected_line)
    tgt = clean(target_line)
    if '共享5V' in exp or '維持5V' in exp or '僅5V' in exp:
        return '5'
    m = re.search(r'回復升壓至\s*(\d+)V', exp)
    if m:
        return m.group(1)
    m = re.search(r'升壓至\s*(\d+)V', exp)
    if m:
        return m.group(1)
    m = re.search(r'VBUS\d*\s*=\s*(\d+)V', tgt)
    if m:
        return m.group(1)
    m = re.search(r'(\d+)V', exp)
    if m:
        return m.group(1)
    return ''


def copy_row_style(ws, src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
        if src._style:
            dst._style = copy.copy(src._style)


def set_row_height(ws, row_idx):
    max_lines = 1
    for c in range(2, 15):
        val = ws.cell(row_idx, c).value
        if isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)
    ws.row_dimensions[row_idx].height = min(15 * max_lines, 90)


def build_measure(step_line, expected_line, send_src, target, gate, request_pdo, protocol, verify, no):
    parts = []
    if step_line:
        parts.append(f'Source Step: {strip_num_prefix(step_line)}')
    if send_src:
        parts.append(f'Send SRC Cap: {strip_num_prefix(send_src)}')
    if target:
        parts.append(f'Target voltage: {strip_num_prefix(target)}')
    if gate:
        parts.append(f'Gate ON: {strip_num_prefix(gate)}')
    if request_pdo:
        parts.append(f'Request PDO: {strip_num_prefix(request_pdo)}')
    if protocol:
        parts.append(f'Protocol: {strip_num_prefix(protocol)}')
    if verify:
        parts.append(f'Verify: {verify}')
    if expected_line:
        parts.append(f'Expected: {strip_num_prefix(expected_line)}')
    parts.append(f'Source Item: 單電源{no}')
    return '\n'.join(parts)


def expand_item(record):
    no = record['no']
    category = record['category']
    test_func = record['test_func']
    test_method = record['test_method']
    expected = record['expected']
    send_src = record['send_src']
    target_v = record['target_v']
    gate = record['gate']
    request_pdo = record['request_pdo']
    protocol = record['protocol']
    verify = record['verify']
    sim = record['sim']
    owner = record['owner']
    due = record['due']

    method_lines = split_lines(test_method)
    expected_lines = split_lines(expected)
    send_lines = split_lines(send_src)
    target_lines = split_lines(target_v)
    gate_lines = split_lines(gate)
    pdo_lines = split_lines(request_pdo)
    proto_lines = split_lines(protocol)

    step_count = max(1, len(expected_lines), len(method_lines))
    group = infer_group(no, test_func)
    items = []

    for idx in range(step_count):
        step_line = method_lines[idx] if idx < len(method_lines) else ''
        expected_line = expected_lines[idx] if idx < len(expected_lines) else ''
        send_line = send_lines[idx] if idx < len(send_lines) else ''
        target_line = target_lines[idx] if idx < len(target_lines) else ''
        gate_line = gate_lines[idx] if idx < len(gate_lines) else ''
        pdo_line = pdo_lines[idx] if idx < len(pdo_lines) else ''
        proto_line = proto_lines[idx] if idx < len(proto_lines) else ''

        typ = infer_typ_voltage(expected_line, target_line)
        desc_core = strip_num_prefix(expected_line) or strip_num_prefix(step_line) or clean(test_func)
        note_parts = []
        if idx == 0:
            note_parts.append(f'Source row/item: 單電源{no}')
            if category:
                note_parts.append(f'Category: {category}')
            if test_func:
                note_parts.append(f'Function: {test_func}')
            if sim:
                note_parts.append(f'sim: {sim}')
            if owner:
                note_parts.append(f'Owner: {owner}')
            if due:
                note_parts.append(f'Due: {due}')

        item = {
            'bin': '2',
            'test_item': group,
            'symbol': f"SP_{no.replace('-', '_')}_S{idx+1}",
            'pwr_seq': 'PWR_VBUS' if no.startswith(('2-', '3-', '4-')) else '',
            'pseudo': '',
            'run': infer_run_pattern(step_line),
            'measure': build_measure(step_line, expected_line, send_line, target_line, gate_line, pdo_line, proto_line, clean(verify), no),
            'desc': f'{no} S{idx+1}: {desc_core}',
            'min': f'=KROW*0.8' if typ else '',
            'typ': typ,
            'max': f'=KROW*1.2' if typ else '',
            'unit': 'V' if typ else '',
            'note': '\n'.join(note_parts),
        }
        items.append(item)
    return items


# load source
src_wb = load_workbook(SRC, read_only=True, data_only=False)
src_ws = src_wb['單電源模式']
records = []
for r in range(16, src_ws.max_row + 1):
    no = clean(src_ws.cell(r, 1).value)
    if not no:
        continue
    verify = clean(src_ws.cell(r, 11).value)
    if should_exclude_verify(verify):
        continue
    records.append({
        'src_row': r,
        'no': no,
        'category': clean(src_ws.cell(r, 2).value),
        'test_func': clean(src_ws.cell(r, 3).value),
        'test_method': src_ws.cell(r, 4).value,
        'expected': src_ws.cell(r, 5).value,
        'send_src': src_ws.cell(r, 6).value,
        'target_v': src_ws.cell(r, 7).value,
        'gate': src_ws.cell(r, 8).value,
        'request_pdo': src_ws.cell(r, 9).value,
        'protocol': src_ws.cell(r, 10).value,
        'verify': verify,
        'sim': clean(src_ws.cell(r, 12).value),
        'owner': clean(src_ws.cell(r, 13).value),
        'due': src_ws.cell(r, 14).value,
    })

expanded = []
for rec in records:
    expanded.extend(expand_item(rec))

# copy base and write workbook
shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']

# clear existing Test Note values from row 2 onward, keep styles
clear_test_note_values(ws, start_row=2, start_col=2, end_col=19)

style_row = 2
start_row = 3
row = start_row
for item in expanded:
    copy_row_style(ws, style_row, row)
    ws.cell(row, COL['bin']).value = item['bin']
    ws.cell(row, COL['test_item']).value = item['test_item']
    ws.cell(row, COL['symbol']).value = item['symbol']
    ws.cell(row, COL['pwr_seq']).value = item['pwr_seq']
    ws.cell(row, COL['pseudo']).value = item['pseudo']
    ws.cell(row, COL['run']).value = item['run']
    ws.cell(row, COL['measure']).value = item['measure']
    ws.cell(row, COL['desc']).value = item['desc']
    ws.cell(row, COL['typ']).value = item['typ']
    if item['typ']:
        ws.cell(row, COL['min']).value = item['min'].replace('ROW', str(row))
        ws.cell(row, COL['max']).value = item['max'].replace('ROW', str(row))
        ws.cell(row, COL['unit']).value = item['unit']
    ws.cell(row, COL['note']).value = item['note']
    set_row_height(ws, row)
    row += 1

# append revision history
if 'Revision history' in wb.sheetnames:
    rev_ws = wb['Revision history']
    rev_row = 3
    while rev_ws.cell(rev_row, 1).value:
        rev_row += 1
    rev_ws.cell(rev_row, 1).value = '2026-04-24'
    rev_ws.cell(rev_row, 2).value = 'R0.1'
    rev_ws.cell(rev_row, 3).value = 'Hermes'
    rev_ws.cell(rev_row, 4).value = 'Generate JD6628H 單電源模式 all items from SA plan into TENJI Test Note'

wb.save(OUT)
print(f'output={OUT}')
print(f'source_items={len(records)}')
print(f'generated_rows={len(expanded)}')
print(f'written_range={start_row}-{row-1}')