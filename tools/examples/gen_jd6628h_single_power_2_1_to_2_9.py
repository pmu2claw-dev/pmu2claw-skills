import copy
import shutil
from datetime import datetime
from openpyxl import load_workbook

SRC = '/home/user/Downloads/JD6628H_Test Plan_20260416.xlsx'
BASE = '/home/user/workspace/tenji-bot/templates/CASE_2026-04-01_JD6628H_BM2922AA_JD6628H_TenjiV2.2_R0.1.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_SinglePwr_2-1_to_2-9_TenjiV2.2_R0.1.xlsm'

COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}


def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def split_lines(v):
    text = clean(v)
    return [x.strip() for x in text.split('\n') if x and x.strip()]


def merged_children(ws):
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                children.add((row, col))
    return children


def clear_test_note_values(ws, start_row, start_col, end_col):
    child_cells = merged_children(ws)
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))


def copy_row_style(ws, src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
        if src._style:
            dst._style = copy.copy(src._style)


def set_row_height(ws, row_idx):
    max_lines = 1
    for c in range(2, 15):
        val = ws.cell(row_idx, c).value
        if isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)
    ws.row_dimensions[row_idx].height = min(15 * max_lines, 90)


def voltage_window(voltage):
    voltage = int(voltage)
    if voltage == 5:
        return ('4.5', '5', '5.5', '4', '6')
    if voltage == 9:
        return ('8.5', '9', '9.5', '8', '10')
    if voltage == 12:
        return ('11.5', '12', '12.5', '11', '13')
    if voltage == 20:
        return ('18.5', '20', '21.5', '19', '21')
    low = max(voltage - 1, 0)
    high = voltage + 1
    return (str(low), str(voltage), str(high), str(low), str(high))


def build_measure(voltage, no, step_idx):
    min_j, typ, max_j, meas_lo, meas_hi = voltage_window(voltage)
    return '\n'.join([
        f'ForceV VBUS1 {typ} 100m',
        f'MeasV VBUS1 {meas_lo} {meas_hi} V_{no.replace("-", "_")}_s{step_idx}',
        f'JudgeV VBUS1 {min_j} {max_j}',
    ])


def infer_group(test_func):
    if '先插C1，再插拔C2' in test_func:
        return 'SinglePwr_C1C2'
    return 'SinglePwr_Mode'


def first_voltage(target_lines):
    first = clean(target_lines[0]) if target_lines else ''
    for token in ('9V', '12V', '20V', '5V'):
        if token in first:
            return int(token[:-1])
    raise ValueError(f'cannot infer first voltage from {target_lines}')


def desc_text(no, step_idx, expected_lines, step_lines):
    if step_idx - 1 < len(expected_lines) and clean(expected_lines[step_idx - 1]):
        return f'{no} S{step_idx}: {clean(expected_lines[step_idx - 1])}'
    return f'{no} S{step_idx}: {clean(step_lines[step_idx - 1])}'


src_wb = load_workbook(SRC, read_only=True, data_only=False)
src_ws = src_wb['單電源模式']
records = []
for r in range(23, 32):
    no = clean(src_ws.cell(r, 1).value)
    if no not in {f'2-{i}' for i in range(1, 10)}:
        continue
    records.append({
        'src_row': r,
        'no': no,
        'category': clean(src_ws.cell(r, 2).value),
        'test_func': clean(src_ws.cell(r, 3).value),
        'test_method': src_ws.cell(r, 4).value,
        'expected': src_ws.cell(r, 5).value,
        'target_v': src_ws.cell(r, 7).value,
        'verify': clean(src_ws.cell(r, 11).value),
        'sim': clean(src_ws.cell(r, 12).value),
        'owner': clean(src_ws.cell(r, 13).value),
        'due': src_ws.cell(r, 14).value,
    })

expanded = []
for rec in records:
    step_lines = split_lines(rec['test_method'])
    expected_lines = split_lines(rec['expected'])
    target_lines = split_lines(rec['target_v'])
    init_v = first_voltage(target_lines)
    share_v = 5
    recover_v = init_v
    group = infer_group(rec['test_func'])
    voltages = [init_v, share_v, recover_v]
    for idx, voltage in enumerate(voltages, start=1):
        note_lines = []
        if idx == 1:
            note_lines.extend([
                f'Source row/item: 單電源{rec["no"]}',
                f'Category: {rec["category"]}',
                f'Function: {rec["test_func"]}',
                f'Verify: {rec["verify"]}',
                f'sim: {rec["sim"]}',
                f'Owner: {rec["owner"]}',
                f'Due: {clean(rec["due"])}',
            ])
        expanded.append({
            'bin': '2',
            'test_item': group,
            'symbol': f'SP_{rec["no"].replace("-", "_")}_S{idx}',
            'pwr_seq': 'PWR_VBUS',
            'pseudo': '',
            'run': '' if idx == 1 else 'Wait 5s',
            'measure': build_measure(voltage, rec['no'], idx),
            'desc': desc_text(rec['no'], idx, expected_lines, step_lines),
            'min': voltage_window(voltage)[0],
            'typ': voltage_window(voltage)[1],
            'max': voltage_window(voltage)[2],
            'unit': 'V',
            'note': '\n'.join([x for x in note_lines if x]),
        })

shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']
clear_test_note_values(ws, start_row=2, start_col=2, end_col=14)
style_row = 2
row = 3
for item in expanded:
    copy_row_style(ws, style_row, row)
    for key in ('bin', 'test_item', 'symbol', 'pwr_seq', 'pseudo', 'run', 'measure', 'desc', 'min', 'typ', 'max', 'unit', 'note'):
        ws.cell(row, COL[key]).value = item[key]
    set_row_height(ws, row)
    row += 1

if 'Revision history' in wb.sheetnames:
    rev_ws = wb['Revision history']
    rev_row = 3
    while rev_ws.cell(rev_row, 1).value:
        rev_row += 1
    rev_ws.cell(rev_row, 1).value = '2026-04-24'
    rev_ws.cell(rev_row, 2).value = 'R0.1'
    rev_ws.cell(rev_row, 3).value = 'Hermes'
    rev_ws.cell(rev_row, 4).value = 'Generate JD6628H 單電源2-1~2-9 into TENJI Test Note workbook'

wb.save(OUT)
print(f'output={OUT}')
print(f'source_items={len(records)}')
print(f'generated_rows={len(expanded)}')
print('range=3-%d' % (row - 1))
