import copy
import shutil
from openpyxl import load_workbook

BASE = '/home/user/workspace-tenji/JD6628H_SinglePwr_2-7_DRAFT_TenjiV2.2_R0.3_openpyxl.xlsm'
OUT = '/home/user/workspace-tenji/JD6628H_SinglePwr_2-7_to_2-9_TenjiV2.2_R0.1.xlsm'

CASES = [
    {
        'case_no': '2-7',
        'function': '先插C1，再插拔C2，C1 Sink PDO > C2 Sink PDO',
        'test_method': '1. C1插入 SINK_20V\n2. 等候5秒後，C2插入 SINK_9V\n3. 等候5秒後，C2拔除\n4. 等候5秒後，C1拔除',
        'expected': [
            '2-7\n1.C1單插可升壓至20V (P1)',
            '2.C2插入時，C1降為5V (P2)，與C2共享5V充電 (P4)',
            '3.C2拔除後，C1回復升壓至20V (P1)',
        ],
        'request_pdo': '1.C1=PDO5\n2.C1=PDO1, C2=PDO1\n3.C1=PDO5',
    },
    {
        'case_no': '2-8',
        'function': '先插C1，再插拔C2，C1 Sink PDO > C2 Sink PDO',
        'test_method': '1. C1插入 SINK_20V\n2. 等候5秒後，C2插入 SINK_12V\n3. 等候5秒後，C2拔除\n4. 等候5秒後，C1拔除',
        'expected': [
            '2-8\n1.C1單插可升壓至20V (P1)',
            '2.C2插入時，C1降為5V (P2)，與C2共享5V充電 (P4)',
            '3.C2拔除後，C1回復升壓至20V (P1)',
        ],
        'request_pdo': '1.C1=PDO5\n2.C1=PDO1, C2=PDO1\n3.C1=PDO5',
    },
    {
        'case_no': '2-9',
        'function': '先插C1，再插拔C2，C1 Sink PDO = C2 Sink PDO',
        'test_method': '1. C1插入 SINK_20V\n2. 等候5秒後，C2插入 SINK_20V\n3. 等候5秒後，C2拔除\n4. 等候5秒後，C1拔除',
        'expected': [
            '2-9\n1.C1單插可升壓至20V (P1)',
            '2.C2插入時，C1降為5V (P2)，與C2共享5V充電 (P4)',
            '3.C2拔除後，C1回復升壓至20V (P1)',
        ],
        'request_pdo': '1.C1=PDO5\n2.C1=PDO1, C2=PDO1\n3.C1=PDO5',
    },
]

COL = {
    'bin': 2,
    'test_item': 3,
    'symbol': 4,
    'pwr_seq': 5,
    'pseudo': 6,
    'run': 7,
    'measure': 8,
    'desc': 9,
    'min': 10,
    'typ': 11,
    'max': 12,
    'unit': 13,
    'note': 14,
}

shutil.copy2(BASE, OUT)
wb = load_workbook(OUT, keep_vba=True)
ws = wb['Test Note']


def merged_children():
    children = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                if rr == min_row and cc == min_col:
                    continue
                children.add((rr, cc))
    return children


def clear_body(start_row=3, start_col=2, end_col=14):
    child_cells = merged_children()
    for r in range(start_row, max(ws.max_row, 40) + 1):
        for c in range(start_col, end_col + 1):
            if (r, c) in child_cells:
                continue
            ws.cell(r, c).value = None
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_col <= end_col and merged_range.max_col >= start_col:
            ws.unmerge_cells(str(merged_range))


# Keep workbook-scoped OS rows only; clear former testcase body first.
clear_body()


def copy_row_style(src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src._style:
            dst._style = copy.copy(src._style)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def note_text(case):
    return '\n'.join([
        f'Source row/item: 單電源{case["case_no"]}',
        'Category: 空載插拔測試 (C1 + C2)',
        f'Function: {case["function"]}',
        f'Test method: {case["test_method"]}',
        'Expected: 1.C1單插可升壓至20V (P1)\\n2.C2插入時，C1降為5V (P2)，與C2共享5V充電 (P4)\\n3.C2拔除後，C1回復升壓至20V (P1)',
        'Target voltage: 1.VBUS1=20V\\n2.VBUS1=5V\\n3.VBUS1=20V',
        f'Request PDO: {case["request_pdo"]}',
        'Verify: co-sim(DRD)',
        'Mapping: workbook-scoped OS_test + one TENJI row per observable step',
    ])

row = 3
for idx, case in enumerate(CASES):
    step1, step2, step3 = row, row + 1, row + 2
    copy_row_style(3, step1)
    copy_row_style(4, step2)
    copy_row_style(5, step3)

    common_note = note_text(case)

    ws.cell(step1, COL['bin']).value = 5
    ws.cell(step1, COL['test_item']).value = 'Single_power1'
    ws.cell(step1, COL['symbol']).value = f"{case['case_no'].replace('-', '_')}_C1_20V"
    ws.cell(step1, COL['pwr_seq']).value = 'PWR_VBUS'
    ws.cell(step1, COL['measure']).value = 'ForceV VIN 5 200mA \nUR k0 ON\nForceI VBUS1 0 5 \nForceV vatach1 7V 0A\nUR k7 ON\nUR k3 ON\nWait 1ms\nPD_command PD_INIT\nWait 2ms\nPD_command PD_req20v\nWait 1ms \nForceI VBUSC_C1 0 5 \nMeasV VBUSC_C1 16 24 - \nJudgeV VBUSC_C1 16 24'
    ws.cell(step1, COL['desc']).value = case['expected'][0]
    ws.cell(step1, COL['min']).value = f'=K{step1}*0.8'
    ws.cell(step1, COL['typ']).value = 20
    ws.cell(step1, COL['max']).value = f'=K{step1}*1.2'
    ws.cell(step1, COL['unit']).value = 'V'
    ws.cell(step1, COL['note']).value = common_note

    ws.cell(step2, COL['bin']).value = 5
    ws.cell(step2, COL['test_item']).value = None
    ws.cell(step2, COL['symbol']).value = f"{case['case_no'].replace('-', '_')}_C2_Attach"
    ws.cell(step2, COL['pwr_seq']).value = None
    ws.cell(step2, COL['measure']).value = 'ForceV vatach2 7V 0A\nPD_command PD_INIT\nWait 2ms\nPD_command PD_req5v\nForceI VBUSC_C1 0 5 \nMeasV VBUSC_C1 4 6 - \nJudgeV VBUSC_C1 4 6'
    ws.cell(step2, COL['desc']).value = case['expected'][1]
    ws.cell(step2, COL['min']).value = f'=K{step2}*0.8'
    ws.cell(step2, COL['typ']).value = 5
    ws.cell(step2, COL['max']).value = f'=K{step2}*1.2'
    ws.cell(step2, COL['unit']).value = 'V'
    ws.cell(step2, COL['note']).value = common_note

    ws.cell(step3, COL['bin']).value = 5
    ws.cell(step3, COL['test_item']).value = None
    ws.cell(step3, COL['symbol']).value = f"{case['case_no'].replace('-', '_')}_C2_detach"
    ws.cell(step3, COL['pwr_seq']).value = None
    ws.cell(step3, COL['measure']).value = 'ForceV vatach2 0V 0A\nWait 1ms\nPD_command PD_INIT\nWait 2ms\nPD_command PD_req20v\nForceI VBUSC_C1 0 5 \nMeasV VBUSC_C1 16 24 - \nJudgeV VBUSC_C1 16 24\nForceV vatach1 0V 0A\nUR k3 OFF\nUR k0 OFF\nForceV VIN 0 200mA \n'
    ws.cell(step3, COL['desc']).value = case['expected'][2]
    ws.cell(step3, COL['min']).value = f'=K{step3}*0.8'
    ws.cell(step3, COL['typ']).value = 20
    ws.cell(step3, COL['max']).value = f'=K{step3}*1.2'
    ws.cell(step3, COL['unit']).value = 'V'
    ws.cell(step3, COL['note']).value = common_note

    row += 3

copy_row_style(6, row)
ws.cell(row, 1).value = None
ws.cell(row, 2).value = 12
ws.cell(row, 3).value = 'OS_test2'
ws.cell(row, 4).value = 'OS_test2'
ws.cell(row, 5).value = 'PWR_OS'
ws.cell(row, 6).value = None
ws.cell(row, 7).value = None
ws.cell(row, 8).value = None
ws.cell(row, 9).value = 'open and short test'
ws.cell(row, 10).value = -0.3
ws.cell(row, 11).value = None
ws.cell(row, 12).value = -0.8
ws.cell(row, 13).value = 'V'
ws.cell(row, 14).value = None

# Clear any leftover rows after tail OS row.
child_cells = merged_children()
for r in range(row + 1, max(ws.max_row, row + 10) + 1):
    for c in range(1, 15):
        if (r, c) in child_cells:
            continue
        ws.cell(r, c).value = None

wb.save(OUT)
print(OUT)
print(f'tail_row={row}')
