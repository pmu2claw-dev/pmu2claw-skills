from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import tempfile
import zipfile
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def load_json(path: str | Path) -> dict[str, Any]:
    return json.loads(Path(path).read_text(encoding='utf-8'))


def dump_json(path: str | Path, payload: dict[str, Any]) -> None:
    out = Path(path)
    ensure_parent(out)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding='utf-8')


def sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with Path(path).open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def vba_sha256(path: str | Path) -> str | None:
    with zipfile.ZipFile(path) as zf:
        try:
            data = zf.read('xl/vbaProject.bin')
        except KeyError:
            return None
    return hashlib.sha256(data).hexdigest()


def copy_cell(src, dst) -> None:
    if isinstance(src, MergedCell):
        return
    dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def rewrite_krow_formula(value: Any, dst_row: int) -> Any:
    if not isinstance(value, str) or not value.startswith('='):
        return value
    return value.replace('KROW', f'K{dst_row}')


def clear_values(ws, start_row: int, end_row: int, start_col: int = 2, end_col: int = 14) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def visible_rows(ws, start_row: int = 1, end_row: int | None = None) -> list[dict[str, Any]]:
    if end_row is None:
        end_row = ws.max_row
    rows: list[dict[str, Any]] = []
    for row in range(start_row, end_row + 1):
        values = [ws.cell(row, col).value for col in range(2, 15)]
        if any(v not in (None, '') for v in values):
            rows.append({'row': row, 'B_to_N': values})
    return rows


def libreoffice_probe(workbook_path: str | Path, outdir: str | Path | None = None) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    temp_dir = None
    target_dir = Path(outdir) if outdir else None
    if target_dir is None:
        temp_dir = Path(tempfile.mkdtemp(prefix='tenji_lo_probe_'))
        target_dir = temp_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    try:
        proc = subprocess.run(
            [
                '/usr/bin/libreoffice',
                '--headless',
                '--convert-to',
                'xlsx',
                '--outdir',
                str(target_dir),
                str(workbook_path),
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        converted = target_dir / f'{workbook_path.stem}.xlsx'
        return {
            'exit_code': proc.returncode,
            'stdout': proc.stdout,
            'stderr': proc.stderr,
            'converted_path': str(converted),
            'exists': converted.exists(),
        }
    finally:
        if temp_dir is not None:
            shutil.rmtree(temp_dir, ignore_errors=True)


def load_workbook_sheet(path: str | Path, sheet_name: str):
    wb = load_workbook(path, keep_vba=True, data_only=False)
    return wb, wb[sheet_name]
