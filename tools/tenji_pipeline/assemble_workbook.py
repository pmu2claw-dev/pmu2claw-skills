from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from common import clear_values, copy_cell, dump_json, ensure_parent, load_json, rewrite_krow_formula, vba_sha256


HEAD_DEFAULT_ROW = 2
TAIL_DEFAULT_ROW = 6


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Assemble validated item specs into a TENJI workbook.')
    parser.add_argument('--template', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--sheet-name', default='Test Note')
    parser.add_argument('--item-spec', action='append', required=True)
    parser.add_argument('--head-donor-workbook', default=None)
    parser.add_argument('--head-donor-row', type=int, default=HEAD_DEFAULT_ROW)
    parser.add_argument('--tail-donor-workbook', default=None)
    parser.add_argument('--tail-donor-row', type=int, default=TAIL_DEFAULT_ROW)
    parser.add_argument('--clear-start-row', type=int, default=2)
    parser.add_argument('--clear-end-row', type=int, default=200)
    parser.add_argument('--report', default=None)
    return parser.parse_args()


def load_specs(paths: list[str]) -> list[dict]:
    return [load_json(path) for path in paths]


def write_row_from_spec(ws, row_idx: int, row_spec: dict) -> None:
    column_map = {
        2: row_spec.get('bin_value', 5),
        3: row_spec.get('test_item_anchor'),
        4: row_spec.get('symbol'),
        5: row_spec.get('pwr_sequence'),
        6: row_spec.get('pseudo_code'),
        7: row_spec.get('run_pattern_wait'),
        8: row_spec.get('measure_condition'),
        9: row_spec.get('description'),
        10: rewrite_krow_formula(row_spec.get('min_formula'), row_idx),
        11: row_spec.get('typ'),
        12: rewrite_krow_formula(row_spec.get('max_formula'), row_idx),
        13: row_spec.get('unit'),
        14: row_spec.get('remarks'),
    }
    for col, value in column_map.items():
        ws.cell(row_idx, col).value = value


def apply_style_source(ws, row_idx: int, row_spec: dict) -> None:
    style_source = row_spec.get('style_source') or {}
    workbook = style_source.get('workbook')
    sheet = style_source.get('sheet') or 'Test Note'
    src_row = style_source.get('row')
    if not workbook or not src_row:
        return
    src_wb = load_workbook(workbook, keep_vba=True, data_only=False)
    src_ws = src_wb[sheet]
    for col in range(1, 15):
        copy_cell(src_ws.cell(src_row, col), ws.cell(row_idx, col))


def merge_test_item_column(ws, start_row: int, end_row: int) -> None:
    if end_row > start_row:
        ws.merge_cells(f'C{start_row}:C{end_row}')


def main() -> int:
    args = parse_args()
    specs = load_specs(args.item_spec)
    template = Path(args.template)
    out = Path(args.output)
    ensure_parent(out)

    out_wb = load_workbook(template, keep_vba=True)
    out_ws = out_wb[args.sheet_name]

    head_donor = load_workbook(args.head_donor_workbook or template, keep_vba=True)[args.sheet_name]
    tail_donor = load_workbook(args.tail_donor_workbook or template, keep_vba=True)[args.sheet_name]

    for merged in list(out_ws.merged_cells.ranges):
        out_ws.unmerge_cells(str(merged))
    clear_values(out_ws, args.clear_start_row, args.clear_end_row)

    for col in range(1, 15):
        copy_cell(head_donor.cell(args.head_donor_row, col), out_ws.cell(2, col))

    current_row = 3
    report_items = []
    for spec in specs:
        case_start = current_row
        first = True
        for row_spec in spec['body_rows']:
            apply_style_source(out_ws, current_row, row_spec)
            write_row_from_spec(out_ws, current_row, row_spec)
            if not first:
                out_ws.cell(current_row, 3).value = None
            first = False
            current_row += 1
        merge_test_item_column(out_ws, case_start, current_row - 1)
        report_items.append(
            {
                'case_id': spec['case_id'],
                'start_row': case_start,
                'end_row': current_row - 1,
            }
        )

    for col in range(1, 15):
        copy_cell(tail_donor.cell(args.tail_donor_row, col), out_ws.cell(current_row, col))
    tail_row = current_row
    clear_values(out_ws, tail_row + 1, args.clear_end_row)

    out_wb.save(out)

    report = {
        'output': str(out),
        'item_count': len(specs),
        'tail_row': tail_row,
        'vba_sha256': vba_sha256(out),
        'items': report_items,
    }
    if args.report:
        dump_json(args.report, report)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
