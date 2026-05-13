from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

from common import dump_json
from schema import make_item_spec, make_row_spec, new_item_spec


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Create a TENJI single-item spec JSON.')
    parser.add_argument('--case-id', required=True)
    parser.add_argument('--family', required=True)
    parser.add_argument('--source-sheet', required=True)
    parser.add_argument('--source-row', type=int, default=None)
    parser.add_argument('--row-count', type=int, default=None)
    parser.add_argument('--answer-workbook', default=None)
    parser.add_argument('--sheet-name', default='Test Note')
    parser.add_argument('--lineage-ref', action='append', default=[])
    parser.add_argument('--output', required=True)
    return parser.parse_args()


def case_token(case_id: str) -> str:
    match = re.search(r'(\d+-\d+)', case_id)
    return match.group(1) if match else case_id


def _cell_str(value) -> str:
    return '' if value is None else str(value)


def extract_rows_from_answer_workbook(answer_workbook: str, case_id: str, sheet_name: str) -> list[dict]:
    workbook_path = Path(answer_workbook)
    wb = load_workbook(workbook_path, keep_vba=True, data_only=False)
    ws = wb[sheet_name]
    token = case_token(case_id)

    start_row = None
    for row in range(1, ws.max_row + 1):
        description = _cell_str(ws.cell(row, 9).value)
        if description.startswith(token):
            start_row = row
            break

    if start_row is None:
        raise ValueError(f'could not find case {case_id} in {workbook_path}')

    rows: list[dict] = []
    row_index = 1
    current_row = start_row
    while current_row <= ws.max_row:
        symbol = ws.cell(current_row, 4).value
        test_item = ws.cell(current_row, 3).value
        if symbol in {'OS_test', 'OS_test2'} or test_item == 'OS_test2':
            break
        if symbol in (None, '') and test_item in (None, '') and ws.cell(current_row, 8).value in (None, ''):
            break

        rows.append(
            make_row_spec(
                row_index_within_case=row_index,
                bin_value=ws.cell(current_row, 2).value,
                test_item_anchor=ws.cell(current_row, 3).value,
                symbol=ws.cell(current_row, 4).value,
                pwr_sequence=ws.cell(current_row, 5).value,
                pseudo_code=ws.cell(current_row, 6).value,
                run_pattern_wait=ws.cell(current_row, 7).value,
                measure_condition=ws.cell(current_row, 8).value,
                description=ws.cell(current_row, 9).value,
                min_formula=ws.cell(current_row, 10).value,
                typ=ws.cell(current_row, 11).value,
                max_formula=ws.cell(current_row, 12).value,
                unit=ws.cell(current_row, 13).value,
                remarks=ws.cell(current_row, 14).value,
                style_source_workbook=str(workbook_path),
                style_source_sheet=sheet_name,
                style_source_row=current_row,
            )
        )
        row_index += 1
        current_row += 1

    if not rows:
        raise ValueError(f'found case {case_id} in {workbook_path} but no body rows were extracted')
    return rows


def build_reasoning_note(case_id: str, lineage_refs: list[str]) -> str:
    if len(lineage_refs) > 1:
        sibling = Path(lineage_refs[1]).name
        return (
            f'Exact-case answer workbook exists for {case_id}. '
            f'{sibling} shows same DSL surface and row shape, confirming this family should stay TENJI-DSL, not prose.'
        )
    return f'Exact-case answer workbook exists for {case_id} and was extracted directly into formal item-spec rows.'


def main() -> int:
    args = parse_args()
    out = Path(args.output)

    if args.answer_workbook:
        answer_workbook = str(Path(args.answer_workbook))
        lineage_refs = [answer_workbook, *args.lineage_ref]
        body_rows = extract_rows_from_answer_workbook(answer_workbook, args.case_id, args.sheet_name)
        spec = make_item_spec(
            case_id=args.case_id,
            family=args.family,
            source_sheet=args.source_sheet,
            source_row=args.source_row,
            lineage_refs=lineage_refs,
            body_rows=body_rows,
            status='reviewed',
            reasoning_note=build_reasoning_note(args.case_id, lineage_refs),
        )
        dump_json(out, spec)
        print(f'Wrote extracted item spec: {out}')
        return 0

    if args.row_count is None:
        raise SystemExit('--row-count is required when --answer-workbook is not provided')

    spec = new_item_spec(
        case_id=args.case_id,
        family=args.family,
        source_sheet=args.source_sheet,
        source_row=args.source_row,
        row_count=args.row_count,
        lineage_refs=args.lineage_ref,
    )
    dump_json(out, spec)
    print(f'Wrote starter item spec: {out}')
    print('Next step: fill body_rows[*] with lineage-backed TENJI DSL, then run validate_single_item.py')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
