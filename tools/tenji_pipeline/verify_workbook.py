from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from common import libreoffice_probe, sha256_file, vba_sha256, visible_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Verify a TENJI workbook package and visible Test Note rows.')
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--sheet-name', default='Test Note')
    parser.add_argument('--start-row', type=int, default=1)
    parser.add_argument('--end-row', type=int, default=40)
    parser.add_argument('--libreoffice-outdir', default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook)
    wb = load_workbook(workbook, keep_vba=True, data_only=False)
    ws = wb[args.sheet_name]
    lo = libreoffice_probe(workbook, args.libreoffice_outdir)
    converted_rows = []
    if lo['exists']:
        converted = Path(lo['converted_path'])
        probe_wb = load_workbook(converted, data_only=False)
        probe_ws = probe_wb[args.sheet_name]
        converted_rows = visible_rows(probe_ws, args.start_row, args.end_row)

    report = {
        'workbook': str(workbook),
        'sha256': sha256_file(workbook),
        'vba_sha256': vba_sha256(workbook),
        'merged_ranges': sorted(str(r) for r in ws.merged_cells.ranges),
        'rows': visible_rows(ws, args.start_row, args.end_row),
        'libreoffice_probe': lo,
        'converted_rows': converted_rows,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if lo['exit_code'] == 0 else 1


if __name__ == '__main__':
    raise SystemExit(main())
