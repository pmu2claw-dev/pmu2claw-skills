#!/usr/bin/env python3
"""
generate_excel.py - 複製 .xlsm 模板，保留 VBA + 格式，只填入資料
使用方式:
  python3 generate_excel.py generate <spec.json> <output.xlsm>
  python3 generate_excel.py append   <spec.json> <existing.xlsm>
"""

import sys
import json
import shutil
import os
from pathlib import Path
import openpyxl

from spec_normalizer import normalize_spec

# ─── 模板路徑 ───────────────────────────────────────────────────────────────
TPL_DIR = Path(__file__).parent.parent / "templates"

def find_latest_template():
    """
    優先挑正式模板，避免把 DEMO/CASE 檔誤當成預設模板。
    優先順序：
      1) TPL_*.xlsm / TPL_*.xlsx
      2) CASE_*.xlsm / CASE_*.xlsx
      3) 其他 .xlsm / .xlsx
    同類型內再取最新修改時間。
    """
    def pick(patterns):
        matches = []
        for pattern in patterns:
            matches.extend(TPL_DIR.glob(pattern))
        matches = sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)
        return matches[0] if matches else None

    for patterns in [
        ["TPL_*.xlsm"],
        ["TPL_*.xlsx"],
        ["CASE_*.xlsm"],
        ["CASE_*.xlsx"],
        ["*.xlsm"],
        ["*.xlsx"],
    ]:
        chosen = pick(patterns)
        if chosen:
            return str(chosen)
    return None

# ─── 安全寫入（只寫值，不動任何格式） ────────────────────────────────────
def copy_row_format(ws, src_row, dst_row):
    """
    將 src_row 整列的格式（框線、字型、填充、對齊、數字格式）複製到 dst_row。
    只複製有資料或有格式的欄，避免覆寫空白欄。
    """
    import copy
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font          = copy.copy(src.font)
            dst.border        = copy.copy(src.border)
            dst.fill          = copy.copy(src.fill)
            dst.alignment     = copy.copy(src.alignment)
            dst.number_format = src.number_format


def auto_fit_row_height(ws):
    """
    清除 row height 的固定值，讓 Excel 開啟後可依內容自動決定行高。

    舊版會對多行內容直接寫死高度，結果 customHeight=True，
    Excel 就不會再自動調整。這裡改成：
    - 若儲存格內容含換行，確保 wrap_text=True
    - 所有有內容的列一律清掉固定高度（height=None）

    註：Excel 是否立即重算顯示，仍取決於客戶端；但只要不把高度寫死，
    就不會卡在固定列高。
    """
    import copy

    for row in ws.iter_rows():
        row_idx = row[0].row
        has_value = False

        for cell in row:
            if cell.value not in (None, ""):
                has_value = True
            if isinstance(cell.value, str) and "\n" in cell.value:
                alignment = copy.copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment

        if has_value:
            # 清掉固定列高，讓 Excel 自行決定
            ws.row_dimensions[row_idx].height = None


def wv(ws, row, col, value):
    """Write Value only — 不碰 fill / font / border / alignment"""
    from openpyxl.utils import get_column_letter
    coord = f"{get_column_letter(col)}{row}"
    # 若目標格在合併區域內，找到主格並寫入
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    ws.cell(row, col).value = value

# ─── Sheet 資料起始行（不含標題行） ─────────────────────────────────────────
# 偵測方式：找第一個完全空白的 row（從 row 2 往下）
def find_data_start_row(ws, header_rows=1):
    """回傳資料區起始 row（跳過 header_rows 行）"""
    return header_rows + 1

def find_next_empty_row(ws, start_row, key_col=2):
    """從 start_row 找第一個 key_col 欄為空的 row"""
    r = start_row
    while r <= ws.max_row:
        if ws.cell(r, key_col).value is None:
            return r
        r += 1
    return r  # max_row + 1

# ─── Test Note 欄位對應 ───────────────────────────────────────────────────
# 依模板 Row 1 標頭：B=Bin C=Test Item D=Symbol E=PWR Seq F=Pseudo code
# G=Run Pattern H=Measure Condition I=Description J=Min K=Typ L=Max M=Unit N=備註
# P=DatasheetP Q=DatasheetS R=QC MIN S=QC MAX
COL_TN = {
    "bin":          2,   # B
    "test_item":    3,   # C
    "symbol":       4,   # D
    "pwr_seq":      5,   # E
    "pseudo_code":  6,   # F
    "run_pattern":  7,   # G
    "measure_cond": 8,   # H
    "description":  9,   # I
    "min":          10,  # J
    "typ":          11,  # K
    "max":          12,  # L
    "unit":         13,  # M
    "note":         14,  # N
    "ds_page":      16,  # P
    "ds_sec":       17,  # Q
    "qc_min":       18,  # R
    "qc_max":       19,  # S
}

# ─── PinMap 欄位對應 ────────────────────────────────────────────────────
# A=Power Monitor B=Pin name C=MVI pin name D=MVI CH E=MCPMU CH
# F=對應PE G=PE Name H=PE CH I=UR Num J=UR 0 K=UR 1 L=上Diode M=下Diode
COL_PM = {
    "power_monitor": 1,
    "pin_name":      2,
    "mvi_pin":       3,
    "mvi_ch":        4,
    "mcpmu_ch":      5,
    "pe_ref":        6,
    "pe_name":       7,
    "pe_ch":         8,
    "ur_num":        9,
    "ur0":           10,
    "ur1":           11,
    "up_diode":      12,
    "dn_diode":      13,
}

# ─── Power sequence 欄位對應 ────────────────────────────────────────────
# 每組 4 欄：Name, 電壓, delay, 最大電壓
PWR_SEQ_GROUPS = {
    "PWR_OS":        1,
    "PWR_Normal":    6,
    "PWR_Leak_Test": 11,
    "PWR_T":         16,
    "PWR_MTP":       21,
    "PWR_MTP_5D5":   26,
    "PWR_MTP_2D5":   31,
}

# ─── 把 commands[] 格式轉成 Test Note 欄位 ────────────────────────────
def commands_to_row(item):
    """
    輸入：{"name": "IQ_VIN_Normal", "commands": [{"cmd":"ForceV","pin":"VIN","params":"1.8V 50mA"}, ...]}
    輸出：Test Note 欄位 dict，可直接給 write_test_note 使用

    ★ 正確的 Test Note 欄位對應（依 TENJI 手冊 REV 0.2.11 & 真實 .xlsm 模板）：
      E = pwr_seq       ← PWR macro 名稱（如 PWR_OS / PWR_Normal），來自 Power Sequence sheet
      F = pseudo_code   ← PseudoCode/SlaveID 宣告（平常空白或 I2C slave address）
      G = run_pattern   ← RunPattern / Wait / AC_pin 等
      H = measure_cond  ← 所有 TENJI 指令全部放這裡（ForceV/I、MeasV/I、UR、
                           JudgeV/I、TuneV/I、TuneReg、Formula、Wait...），
                           每條指令各一行（\n 分隔）

    Note：pwr_seq 欄（E）只放 PWR macro 名稱，不是 ForceV/ForceI 指令。
          ForceV/ForceI 等所有 TENJI 指令一律放 measure_cond（H 欄）。
    """
    name    = item.get("name", item.get("test_item", ""))
    cmds    = item.get("commands", [])

    meas_parts   = []   # H 欄：所有 TENJI 指令
    judge_min    = ""
    judge_max    = ""
    judge_unit   = ""

    JUDGE_CMDS = ("JudgeV", "JudgeI", "JudgeFreq", "JudgeDuty", "JudgeDbl",
                  "JudgeVP", "JudgeIP")

    for c in cmds:
        cmd    = c.get("cmd", "").strip()
        pin    = c.get("pin", "").strip()
        params = c.get("params", "").strip()

        if not cmd:
            continue

        # 所有 TENJI 指令一律放 measure_cond（H欄），多行
        parts = [cmd]
        if pin: parts.append(pin)
        if params: parts.append(params)
        meas_parts.append(" ".join(parts))

        # 從 Judge 指令解析 min/max/unit（用於 I/J/K/L 欄）
        if any(cmd.startswith(j) for j in JUDGE_CMDS):
            tokens = params.split() if params else []
            if len(tokens) >= 2:
                judge_min = tokens[0]
                judge_max = tokens[1]
            if len(tokens) >= 3:
                judge_unit = tokens[2]
            if not judge_unit and judge_min:
                for unit in ("uA","mA","A","uV","mV","V","MHz","kHz","Hz","us","ms","s"):
                    if judge_min.endswith(unit) or judge_max.endswith(unit):
                        judge_unit = unit
                        break

    measure_cond = item.get("measure_cond") or "\n".join(meas_parts)
    min_value = item.get("min") or judge_min
    max_value = item.get("max") or judge_max
    unit_value = item.get("unit") or judge_unit

    return {
        "bin":          item.get("bin", ""),
        "test_item":    name,
        "symbol":       item.get("symbol", name),
        "pwr_seq":      item.get("pwr_seq", ""),        # E欄：PWR macro 名稱
        "pseudo_code":  item.get("pseudo_code", ""),    # F欄：SlaveID 宣告
        "run_pattern":  item.get("run_pattern", ""),    # G欄：RunPattern/Wait
        "measure_cond": measure_cond,                     # H欄：所有 TENJI 指令
        "description":  item.get("description", ""),
        "min":          min_value,
        "typ":          item.get("typ", ""),
        "max":          max_value,
        "unit":         unit_value,
        "note":         item.get("note", ""),
        "ds_page":      item.get("ds_page", ""),
        "ds_sec":       item.get("ds_sec", ""),
        "qc_min":       item.get("qc_min", ""),
        "qc_max":       item.get("qc_max", ""),
    }

# ─── 寫入 Test Note ──────────────────────────────────────────────────────
def write_test_note(ws, spec):
    start = find_data_start_row(ws, header_rows=1)
    # 找到第一個空白行（key=Test Item 欄 col 3）
    row = find_next_empty_row(ws, start, key_col=3)
    raw_items = spec.get("test_items", [])
    for raw_item in raw_items:
        # 從第 3 行開始，複製前一列整列格式（框線、字型、填充、對齊）
        if row >= 3:
            copy_row_format(ws, row - 1, row)
        # 若 item 帶有 commands[] 就先轉換，否則直接用原有欄位
        item = commands_to_row(raw_item) if "commands" in raw_item else raw_item
        wv(ws, row, COL_TN["bin"],          item.get("bin", ""))
        wv(ws, row, COL_TN["test_item"],    item.get("test_item", ""))
        wv(ws, row, COL_TN["symbol"],       item.get("symbol", ""))
        wv(ws, row, COL_TN["pwr_seq"],      item.get("pwr_seq", ""))
        wv(ws, row, COL_TN["pseudo_code"],  item.get("pseudo_code", ""))
        wv(ws, row, COL_TN["run_pattern"],  item.get("run_pattern", ""))
        wv(ws, row, COL_TN["measure_cond"], item.get("measure_cond", ""))
        wv(ws, row, COL_TN["description"],  item.get("description", ""))
        wv(ws, row, COL_TN["min"],          item.get("min", ""))
        wv(ws, row, COL_TN["typ"],          item.get("typ", ""))
        wv(ws, row, COL_TN["max"],          item.get("max", ""))
        wv(ws, row, COL_TN["unit"],         item.get("unit", ""))
        wv(ws, row, COL_TN["note"],         item.get("note", ""))
        wv(ws, row, COL_TN["ds_page"],      item.get("ds_page", ""))
        wv(ws, row, COL_TN["ds_sec"],       item.get("ds_sec", ""))
        wv(ws, row, COL_TN["qc_min"],       item.get("qc_min", ""))
        wv(ws, row, COL_TN["qc_max"],       item.get("qc_max", ""))
        row += 1
    print(f"  [Test Note] 寫入 {len(raw_items)} 測項，從 row {start} 開始")

# ─── 寫入 PinMap ─────────────────────────────────────────────────────────
def write_pinmap(ws, spec):
    start = find_data_start_row(ws, header_rows=1)
    row = find_next_empty_row(ws, start, key_col=2)
    pins = spec.get("pinmap", spec.get("pins", []))
    for pin in pins:
        wv(ws, row, COL_PM["power_monitor"], pin.get("power_monitor", ""))
        wv(ws, row, COL_PM["pin_name"],      pin.get("pin_name", pin.get("pin", "")))
        wv(ws, row, COL_PM["mvi_pin"],       pin.get("mvi_pin", ""))
        wv(ws, row, COL_PM["mvi_ch"],        pin.get("mvi_ch", ""))
        wv(ws, row, COL_PM["mcpmu_ch"],      pin.get("mcpmu_ch", ""))
        wv(ws, row, COL_PM["pe_ref"],        pin.get("pe_ref", ""))
        wv(ws, row, COL_PM["pe_name"],       pin.get("pe_name", ""))
        wv(ws, row, COL_PM["pe_ch"],         pin.get("pe_ch", ""))
        wv(ws, row, COL_PM["ur_num"],        pin.get("ur_num", ""))
        wv(ws, row, COL_PM["ur0"],           pin.get("ur0", ""))
        wv(ws, row, COL_PM["ur1"],           pin.get("ur1", ""))
        wv(ws, row, COL_PM["up_diode"],      pin.get("up_diode", ""))
        wv(ws, row, COL_PM["dn_diode"],      pin.get("dn_diode", ""))
        row += 1
    print(f"  [PinMap] 寫入 {len(pins)} pin，從 row {start} 開始")

# ─── 寫入 Power sequence ─────────────────────────────────────────────────
def write_power_sequence(ws, spec):
    pwr_raw = spec.get("power_sequence", {})

    # 相容兩種格式：
    # 1. dict 格式：{"PWR_Normal": [...], "PWR_OS": [...]}
    # 2. list 格式：[{"pin": "VDD", "action": "ForceV 5V", "delay": "1ms"}, ...]
    #    → list 格式全部寫入 PWR_Normal 欄
    if isinstance(pwr_raw, list):
        pwr_data = {"PWR_Normal": pwr_raw}
    else:
        pwr_data = pwr_raw

    for grp_name, base_col in PWR_SEQ_GROUPS.items():
        entries = pwr_data.get(grp_name, [])
        if not entries:
            continue
        row = find_next_empty_row(ws, 2, key_col=base_col)
        for entry in entries:
            # entry 支援 dict 格式：
            #   {"pin": "VDD", "voltage": 3.3, "delay": "1ms", "max_v": 3.6}
            #   或 {"pin": "PWR_VBUS", "action": "ForceV 5V 500mA", "delay": "1ms"}
            pin_name = entry.get("pin", "")
            voltage  = entry.get("voltage", entry.get("action", ""))
            delay    = entry.get("delay", "")
            max_v    = entry.get("max_v", "")
            wv(ws, row, base_col,     pin_name)
            wv(ws, row, base_col + 1, voltage)
            wv(ws, row, base_col + 2, delay)
            wv(ws, row, base_col + 3, max_v)
            row += 1
    print(f"  [Power sequence] 完成")

# ─── 寫入 Revision history ───────────────────────────────────────────────
def write_revision(ws, spec):
    # Row 1 = "Revision History" header, Row 2 = 欄位頭 (Date/Revision/Reviser/Change Item)
    # Row 3 開始是資料（模板已有 R0.0 那筆）
    start = 3
    row = find_next_empty_row(ws, start, key_col=1)
    for rev in spec.get("revisions", []):
        wv(ws, row, 1, rev.get("date", ""))
        wv(ws, row, 2, rev.get("revision", ""))
        wv(ws, row, 3, rev.get("reviser", ""))
        wv(ws, row, 4, rev.get("change", ""))
        row += 1
    print(f"  [Revision history] 完成")

# ─── 安全備份機制 ────────────────────────────────────────────────────────
def create_backup(file_path):
    """在修改前建立帶有時間戳記的備份"""
    if not os.path.exists(file_path):
        return
    import datetime
    p = Path(file_path)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = p.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / f"{p.stem}_{ts}{p.suffix}"
    try:
        shutil.copy2(file_path, backup_path)
        print(f"  [Backup] 已建立備份：{backup_path.name}")
    except Exception as e:
        print(f"  [Backup] 備份失敗：{e}")

# ─── 主流程 ──────────────────────────────────────────────────────────────
def generate(spec, output_path):
    """複製 .xlsm 模板 → keep_vba=True → 只填資料 → 存 .xlsm"""
    spec = normalize_spec(spec)
    create_backup(output_path)
    tpl = find_latest_template()
    if not output_path.endswith(".xlsm"):
        output_path = output_path.rsplit(".", 1)[0] + ".xlsm"

    if tpl:
        print(f"[generate_excel] 使用模板：{Path(tpl).name}")
        shutil.copy2(tpl, output_path)
        wb = openpyxl.load_workbook(output_path, keep_vba=True)
    else:
        print("[generate_excel] 無模板，從零建立")
        wb = openpyxl.Workbook()
        for sname in ["Test Note", "PinMap", "Power sequence", "PseudoCode", "Revision history"]:
            if sname not in wb.sheetnames:
                wb.create_sheet(sname)

    if "Test Note" in wb.sheetnames:
        write_test_note(wb["Test Note"], spec)
    if "PinMap" in wb.sheetnames:
        write_pinmap(wb["PinMap"], spec)
    if "Power sequence" in wb.sheetnames:
        write_power_sequence(wb["Power sequence"], spec)
    if "Revision history" in wb.sheetnames:
        write_revision(wb["Revision history"], spec)

    # 自動調整行高
    for sname in wb.sheetnames:
        auto_fit_row_height(wb[sname])

    wb.save(output_path)
    print(f"[generate_excel] ✅ 輸出：{output_path}")
    return output_path


def append(spec, target_path):
    """追加資料到現有 .xlsm（不複製模板，直接開現有檔案）"""
    spec = normalize_spec(spec, existing_excel=target_path)
    if not os.path.exists(target_path):
        print(f"[generate_excel] 找不到 {target_path}，改為 generate")
        return generate(spec, target_path)

    create_backup(target_path)
    print(f"[generate_excel] 追加到：{target_path}")
    wb = openpyxl.load_workbook(target_path, keep_vba=True)

    if "Test Note" in wb.sheetnames:
        write_test_note(wb["Test Note"], spec)
    if "PinMap" in wb.sheetnames:
        write_pinmap(wb["PinMap"], spec)
    if "Power sequence" in wb.sheetnames:
        write_power_sequence(wb["Power sequence"], spec)
    if "Revision history" in wb.sheetnames:
        write_revision(wb["Revision history"], spec)

    # 自動調整行高
    for sname in wb.sheetnames:
        auto_fit_row_height(wb[sname])

    wb.save(target_path)
    print(f"[generate_excel] ✅ 追加完成：{target_path}")
    return target_path


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法: python3 generate_excel.py <generate|append> <spec.json> <output.xlsm>")
        sys.exit(1)

    action      = sys.argv[1]
    spec_path   = sys.argv[2]
    output_path = sys.argv[3]

    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    if action == "generate":
        generate(spec, output_path)
    elif action == "append":
        append(spec, output_path)
    else:
        print(f"未知 action: {action}")
        sys.exit(1)
