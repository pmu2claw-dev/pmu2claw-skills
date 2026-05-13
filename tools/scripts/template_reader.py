#!/usr/bin/env python3
"""
template_reader.py
讀取 TENJI Excel 模板，輸出結構化摘要給 AI 參考
用法：python3 template_reader.py <excel_path>
"""

import sys
import json
import openpyxl
from pathlib import Path


def safe_cell(ws, row, col):
    """安全讀取儲存格，返回字串"""
    try:
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None else ""
    except Exception:
        return ""


def read_template(path: str) -> dict:
    """讀取 TENJI Excel，回傳結構摘要"""
    result = {
        "path": path,
        "filename": Path(path).name,
        "sheets": [],
        "pinmap": [],
        "test_items": [],
        "column_headers": {},
        "sample_rows": {},
    }

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        result["error"] = str(e)
        return result

    result["sheets"] = wb.sheetnames

    for sheet_name in wb.sheetnames:
        if sheet_name not in ["Revision history", "Test Note", "PinMap", "Power sequence", "PseudoCode"]:
             continue
        ws = wb[sheet_name]
        lower_name = sheet_name.lower()

        # 讀取標頭列（前兩列）
        headers = []
        try:
            # 限制只讀前 10 欄，前 3 列，減少大型表格的解析負擔
            for row in ws.iter_rows(min_row=1, max_row=3, max_col=15, values_only=True):
                row_vals = [str(v).strip() if v else "" for v in row]
                if any(row_vals):
                    headers.append(row_vals)
        except Exception:
            pass
        result["column_headers"][sheet_name] = headers

        # 收集樣本資料列（最多 3 行）
        sample = []
        try:
            row_count = 0
            for row in ws.iter_rows(min_row=4, max_row=10, max_col=15, values_only=True):
                row_vals = [str(v).strip() if v else "" for v in row]
                if any(row_vals):
                    sample.append(row_vals)
                    row_count += 1
                    if row_count >= 3:
                        break
        except Exception:
            pass
        result["sample_rows"][sheet_name] = sample

        # PinMap 表（只處理關鍵 Sheet）
        if "pin" in lower_name:
            pins = []
            try:
                for row in ws.iter_rows(min_row=2, max_row=50, max_col=10, values_only=True):
                    if not any(row):
                        continue
                    r = [str(v).strip() if v else "" for v in row]
                    if r[0] and r[0] not in ("Pin", "PIN", "pin", ""):
                        pin_entry = {
                            "pin": r[0] if len(r) > 0 else "",
                            "type": r[1] if len(r) > 1 else "",
                        }
                        if pin_entry["pin"]:
                            pins.append(pin_entry)
                    if len(pins) >= 15:
                        break
            except Exception:
                pass
            result["pinmap"] = pins

        # TestNote 表（測試項目範例）
        if "test" in lower_name or "note" in lower_name:
            items = []
            current_item = None
            try:
                for row in ws.iter_rows(min_row=3, max_row=40, max_col=10, values_only=True):
                    r = [str(v).strip() if v else "" for v in row]
                    if not any(r):
                        continue
                    if r[2] and r[2] not in ("Symbol", "symbol", ""):
                        current_item = {
                            "name": r[2],
                            "commands": []
                        }
                        items.append(current_item)
                        # 簡單抓取該列的指令片段
                        if len(r) > 7 and r[7]:
                            current_item["commands"].append({"cmd": r[7][:30], "pin": "", "params": ""})
                    if len(items) >= 5:
                        break
            except Exception:
                pass
            result["test_items"] = items

    wb.close()
    return result


def summarize_for_ai(template_data: dict) -> str:
    """將模板資料轉成給 AI 的文字摘要"""
    lines = [f"## 模板：{template_data['filename']}"]
    lines.append(f"工作表：{', '.join(template_data.get('sheets', []))}")

    if template_data.get("pinmap"):
        lines.append("\n### PinMap")
        for p in template_data["pinmap"][:15]:
            lines.append(f"  {p['pin']} | {p['type']} | {p['instrument']} | N={p['N']} | O={p['O']}")

    if template_data.get("test_items"):
        lines.append("\n### 測試項目範例")
        for item in template_data["test_items"][:6]:
            lines.append(f"  [{item['name']}]")
            for cmd in item.get("commands", [])[:5]:
                lines.append(f"    {cmd['cmd']} {cmd['pin']} {cmd['params']}")

    if template_data.get("column_headers"):
        lines.append("\n### 欄位標頭")
        for sheet, hdrs in template_data["column_headers"].items():
            if hdrs:
                lines.append(f"  {sheet}: {hdrs[0][:8]}")

    return "\n".join(lines)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "需要提供 Excel 路徑"}))
        sys.exit(1)

    path = sys.argv[1]
    mode = sys.argv[2] if len(sys.argv) > 2 else "summary"

    data = read_template(path)

    if mode == "json":
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        print(summarize_for_ai(data))
